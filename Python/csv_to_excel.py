# --------------------------------------------------------------------------
# Python function: csv_to_excel.py
# Requires the following libraries: pandas and openpyxl
# 
# Purpose: 
# - Converts CSV files in a folder to an Excel spreadsheet.
# - Each CSV is converted to a separate tab.
# - The tab name is based on the CSV filename, but sanitized to avoid invalid names.
# - Data is converted to an Excel table with the same name as the tab name.
# 
# Usage:
# - By default, the script converts all CSV files in the same location as the Python file
#   and saves the Excel file as "combined_csvs.xlsx" in the same folder.
# - The script can be executed via the command line with optional parameters:
#
# Command-line Arguments:
#  --path <directory>       : (Optional) Path to the folder containing CSV files. Defaults to the current folder.
#  --output <filename.xlsx> : (Optional) Name of the output Excel file. Defaults to "combined_csvs.xlsx".
#  --delimiter <char>       : (Optional) Delimiter used in CSV files (e.g. "|"). Defaults to a comma (",").
#  --ignoreCsvs <files>     : (Optional) Comma-separated list of CSV files to ignore (e.g., "file1.csv,file2.csv").
# 
# Example Usage:
# 1. Convert all CSVs in the current directory:
#    python csv_to_excel.py
#
# 2. Specify a custom output filename:
#    python csv_to_excel.py --output my_output.xlsx
#
# 3. Convert CSVs from a specific folder:
#    python csv_to_excel.py --path /path/to/csvs/
#
# 4. Use a different delimiter (e.g., tab-separated files):
#    python csv_to_excel.py --delimiter "\t"
#
# 5. Ignore specific CSV files:
#    python csv_to_excel.py --ignoreCsvs "ignore_this.csv,skip_this_too.csv"
#
# Restrictions:
# - The output file must be stored in the same folder as the CSV files (specified by --path), or in a subfolder or sibling folder.
# - If an invalid location is specified (e.g., an unrelated absolute path), the script will raise an error.
# - If --path and --output contain spaces, they need to be specified in quotes. 
# - The functionality currently does not support multiple or mixed delimiters across csv files.
#
# Assumptions:
# - The options --path and --output can be absolute or relative paths
# - The output filename must end with ".xlsx"; otherwise, an error will be raised.
# - Files listed in --ignoreCsvs must have the .csv extension and must be located within the directory specified by --path.  
# - It is assumed that CSV files are properly structured with headers and consistent columns.
# - If a CSV is malformed (e.g., missing headers or inconsistent columns), unexpected behavior may occur, such as errors or missing data.
# --------------------------------------------------------------------------

import os
import re
import sys
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def verify_libraries(required_libraries: list) -> None:
    missing_libraries = [lib for lib in required_libraries if not __import__(lib, globals(), locals(), [], 0)]
    if missing_libraries:
        print(f"Missing libraries: {', '.join(missing_libraries)}")
        print(f"Install with: pip install {' '.join(missing_libraries)}")
        sys.exit(1)

def detect_delimiter(file_path: str) -> str:
    with open(file_path, 'r', encoding='utf-8') as file:
        first_line = file.readline()
    delimiters = [',', ';', '\t', '|']
    return max(delimiters, key=lambda d: first_line.count(d))

def sanitize_name(name: str, max_length: int = 31) -> str:
    sanitized = re.sub(r"[^\w]", "_", name)
    return sanitized[:max_length] if len(sanitized) <= max_length else sanitized[:max_length-3] + "..."

def validate_path(path: str) -> str:
    abs_path = os.path.abspath(path)
    if not os.path.exists(abs_path) or not os.path.isdir(abs_path):
        raise ValueError(f"Invalid directory path: {abs_path}")
    return abs_path

def validate_output_path(output: str, input_path: str) -> str:
    abs_output = os.path.abspath(output)
    if not abs_output.lower().endswith(".xlsx"):
        raise ValueError("Output file must have a .xlsx extension")
    output_folder = os.path.dirname(abs_output)
    input_abs = os.path.abspath(input_path)
    if not (output_folder.startswith(input_abs) or os.path.dirname(output_folder) == os.path.dirname(input_abs)):
        raise ValueError("Output must be in input directory, its subfolder, or a sibling folder")
    os.makedirs(output_folder, exist_ok=True)
    temp_file = os.path.join(output_folder, ".write_test.tmp")
    try:
        with open(temp_file, "w") as f:
            f.write("test")
        os.remove(temp_file)
    except IOError:
        raise PermissionError(f"Cannot write to output folder: {output_folder}")
    return abs_output

def combine_csvs(path: str = None, output: str = "combined_csvs.xlsx", ignore_csvs: list = []):
    path = validate_path(path or os.getcwd())
    output_path = validate_output_path(output, path)
    wb = Workbook()
    wb.remove(wb.active)
    for filename in os.listdir(path):
        if filename.endswith(".csv") and filename not in ignore_csvs:
            try:
                file_path = os.path.join(path, filename)
                delimiter = detect_delimiter(file_path)
                df = pd.read_csv(file_path, delimiter=delimiter, header=0)
                if df.empty:
                    continue
                tab_name = sanitize_name(os.path.splitext(filename)[0])
                ws = wb.create_sheet(title=tab_name)
                ws.append(list(df.columns))
                for row in df.itertuples(index=False):
                    ws.append(row)
                table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
                table = Table(displayName=tab_name, ref=table_range)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(table)
            except Exception as e:
                print(f"Error processing {filename}: {e}")
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")

if __name__ == "__main__":
    verify_libraries(["pandas", "openpyxl"])
    parser = argparse.ArgumentParser(description="Combine CSV files into a single Excel file.")
    parser.add_argument("--path", type=str, help="Directory containing CSV files.")
    parser.add_argument("--output", type=str, default="combined_csvs.xlsx", help="Output Excel filename.")
    parser.add_argument("--ignoreCsvs", type=str, help="Comma-separated list of CSV files to ignore.")
    args = parser.parse_args()
    ignore_csvs = args.ignoreCsvs.split(",") if args.ignoreCsvs else []
    combine_csvs(path=args.path, output=args.output, ignore_csvs=ignore_csvs)
