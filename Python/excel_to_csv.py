# --------------------------------------------------------------------------
# Python function: excel_to_csv
# Requires the following libraries: pandas and openpyxl
# 
# Purpose: 
# - Extracts data from each sheet in an Excel workbook and saves them as CSV files.
# - If a sheet contains named tables, each table is saved as a separate CSV.
# - Otherwise, the entire sheet is saved as a CSV.
#
# Usage:
# - The script reads an Excel file (.xlsx) and extracts data into CSV files.
# - By default, CSV files are saved in the same directory as the Excel file.
# - The script can be executed via the command line with optional parameters:
#
# Command-line Arguments:
#  --filename <file.xlsx>   : (Required) Path to the input Excel file.
#  --output <directory>     : (Optional) Output folder for CSV files. Defaults to the Excel file's directory.
#  --delimiter <char>       : (Optional) Delimiter used in CSV files (e.g., ";"). Defaults to a comma (",").
#
# Example Usage:
# 1. Extract all sheets and tables from an Excel file:
#    python excel_to_csv.py
#
# The user will be prompted to provide the filename of the spreadsheet.
#
# 2. Extract all sheets and tables from an Excel file:
#    python excel_to_csv.py --filename my_data.xlsx
#
# 3. Specify an output folder for CSVs:
#    python excel_to_csv.py --filename my_data.xlsx --output /path/to/output/
#
# 4. Use a different delimiter (e.g., semicolon):
#    python excel_to_csv.py --filename my_data.xlsx --delimiter ";"
#
# Restrictions:
# - The output folder must be valid and writable; otherwise, an error will occur.
# - The script does not support multiple or mixed delimiters within a single run.
# - If an Excel sheet contains tables, only the named table areas are extracted, not the entire sheet.
#
# Assumptions:
# - The input Excel file (.xlsx) exists and is accessible.
# - If an output folder is specified, it must be a valid directory.
# - If no tables exist in a sheet, the whole sheet is saved as a CSV with name <sheetname>.csv.
# - If tables exist on a sheet, each table is saved as a separate CS with name <sheetname> - <tablename>.csv
# - Named tables in Excel are well-defined and contain structured data.
# - Empty sheets are skipped (not saved as csv).
# --------------------------------------------------------------------------


import pandas as pd
from openpyxl import load_workbook
import os
import argparse
import sys

def verify_libraries(required_libraries: list) -> None:
    missing_libraries = []
    for library in required_libraries:
        try:
            __import__(library)
        except ImportError:
            missing_libraries.append(library)
    if missing_libraries:
        print(f"The following libraries are missing:", flush=True)
        for lib in missing_libraries:
            print(f" - {lib}", flush=True)
        print(f"Install them with: pip install " + " ".join(missing_libraries), flush=True)
        sys.exit(1)
    print(f"Libraries in place", flush=True)

def extract_sheets_to_csv(excel_file, output_folder=None):
    print(f"Processing file: {excel_file}", flush=True)
    if output_folder is None or output_folder == "":
        output_folder = os.getcwd()
    
    os.makedirs(output_folder, exist_ok=True)
    print(f"Saving CSV files to: {output_folder}", flush=True)

    try:
        wb = load_workbook(excel_file, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}", flush=True)
        return

    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}", flush=True)
        sheet = wb[sheet_name]

        if all(cell.value is None for row in sheet.iter_rows() for cell in row):
            print(f"Skipping empty sheet: {sheet_name}", flush=True)
            continue

        try:
            if sheet.tables:
                for table_name in sheet.tables.keys():
                    table_obj = sheet.tables[table_name]
                    if hasattr(table_obj, 'ref'):
                        table_range = table_obj.ref
                        start_cell, end_cell = table_range.split(":")

                        min_row = sheet[start_cell].row
                        max_row = sheet[end_cell].row
                        min_col = sheet[start_cell].column
                        max_col = sheet[end_cell].column

                        data = [
                            [sheet.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
                            for r in range(min_row, max_row + 1)
                        ]

                        df = pd.DataFrame(data)
                        csv_filename = f"{sheet_name} - {table_name}.csv"
                        df.to_csv(os.path.join(output_folder, csv_filename), index=False, sep=",", header=False)
                        print(f"Saved table: {csv_filename}", flush=True)
            else:
                data = [[cell.value for cell in row] for row in sheet.iter_rows()]
                df = pd.DataFrame(data)
                csv_filename = f"{sheet_name}.csv"
                df.to_csv(os.path.join(output_folder, csv_filename), index=False, sep=",", header=False)
                print(f"Saved sheet: {csv_filename}", flush=True)
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}", flush=True)

if __name__ == "__main__":
    print("Starting script execution...", flush=True)
    print("Verifying required libraries...", flush=True)
    verify_libraries(["pandas", "openpyxl"])

    print("Parsing command-line arguments...", flush=True)
    parser = argparse.ArgumentParser(description="Extract sheets and tables from an Excel file and save them as CSVs.")
    parser.add_argument("--filename", help="Path to the input Excel file")
    parser.add_argument("--output", help="Output folder for CSV files (default: same as Excel file location)")
    args = parser.parse_args()

    if not args.filename:
        print("No filename provided via command-line arguments.", flush=True)
        args.filename = input("Please provide the Excel filename (.xlsx) to be converted: ").strip()
    
    if not args.filename or not args.filename.lower().endswith(".xlsx"):
        print("Error: No valid filename provided. Aborting...", flush=True)
        sys.exit(1)
    
    if not os.path.exists(args.filename):
        print("Error: The specified file does not exist.", flush=True)
        sys.exit(1)
    
    if args.output:
        if not os.path.exists(args.output):
            try:
                os.makedirs(args.output)
            except Exception as e:
                print(f"Error: Unable to create output directory: {e}", flush=True)
                sys.exit(1)
        elif not os.path.isdir(args.output):
            print("Error: The specified output path is not a directory.", flush=True)
            sys.exit(1)
    
    print(f"Filename received: {args.filename}", flush=True)
    print("Starting process...", flush=True)

    extract_sheets_to_csv(args.filename, args.output)
